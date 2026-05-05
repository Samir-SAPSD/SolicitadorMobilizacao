import os
import subprocess
import sys
import tempfile
from flask import Flask, render_template, request, Response, stream_with_context, jsonify, send_from_directory
from werkzeug.utils import secure_filename
from datetime import datetime
import json
import re
import threading
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.json.ensure_ascii = False

# Configurações
UPLOAD_FOLDER = 'templates'
REPORTS_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx'}

ACTIVE_JOBS = 0
JOBS_LOCK = threading.Lock()

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
TEMPLATE_FILENAME = 'ModeloSolicitacaoMob.xlsx'
TEMPLATE_STATUS_FILE = os.path.join('templates', 'template_update_status.json')


def start_job():
    global ACTIVE_JOBS
    with JOBS_LOCK:
        ACTIVE_JOBS += 1


def end_job():
    global ACTIVE_JOBS
    with JOBS_LOCK:
        if ACTIVE_JOBS > 0:
            ACTIVE_JOBS -= 1


def generate_report_excel(report_data: dict) -> tuple[str, str]:
    """
    Gera um arquivo Excel formatado a partir do JSON de relatório emitido pelo PowerShell.
    Retorna (nome_arquivo, caminho_absoluto).
    """
    # ── Cores Vestas ──────────────────────────────────────────────────────────
    COLOR_NIGHT_SKY  = "1F3144"
    COLOR_BLUE_SKY   = "005AFF"
    COLOR_LIGHT_GREY = "E3E5E8"
    COLOR_GREEN_FILL = "D4EDDA"
    COLOR_RED_FILL   = "F8D7DA"
    COLOR_WHITE      = "FFFFFF"
    COLOR_LABEL_BG   = "EBF3FB"

    fill_header    = PatternFill("solid", fgColor=COLOR_NIGHT_SKY)
    fill_col_hdr   = PatternFill("solid", fgColor=COLOR_NIGHT_SKY)
    fill_label     = PatternFill("solid", fgColor=COLOR_LABEL_BG)
    fill_success   = PatternFill("solid", fgColor=COLOR_GREEN_FILL)
    fill_error     = PatternFill("solid", fgColor=COLOR_RED_FILL)
    fill_light     = PatternFill("solid", fgColor=COLOR_LIGHT_GREY)

    font_white_bold  = Font(name="Calibri", bold=True,  color=COLOR_WHITE, size=13)
    font_col_hdr     = Font(name="Calibri", bold=True,  color=COLOR_WHITE, size=10)
    font_label       = Font(name="Calibri", bold=True,  color=COLOR_NIGHT_SKY, size=10)
    font_value       = Font(name="Calibri", bold=False, color=COLOR_NIGHT_SKY, size=10)
    font_value_bold  = Font(name="Calibri", bold=True,  color=COLOR_NIGHT_SKY, size=10)

    thin_border_side = Side(style="thin", color="CCCCCC")
    thin_border      = Border(
        left=thin_border_side, right=thin_border_side,
        top=thin_border_side,  bottom=thin_border_side
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório de Submissão"

    items             = report_data.get("items", [])
    submission_dt     = report_data.get("submission_datetime", "")
    id_mob            = report_data.get("id_mobilizacao", "")
    requester_name    = report_data.get("requester_name", "")
    requester_email   = report_data.get("requester_email", "")

    # ── Coletar todos os display names usados, preservando ordem de aparição ──
    display_cols: list[str] = []
    seen: set[str] = set()
    for item in items:
        for col_name in (item.get("fields") or {}).keys():
            if col_name not in seen:
                seen.add(col_name)
                display_cols.append(col_name)

    total_cols = 2 + len(display_cols)  # "ID SP" + "Status" + campo...
    last_col_letter = get_column_letter(max(total_cols, 3))

    # ── ROW 1 – Título principal ───────────────────────────────────────────────
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws["A1"]
    title_cell.value = "RELATÓRIO DE SUBMISSÃO — MOBILIZAÇÕES VESTAS"
    title_cell.fill  = fill_header
    title_cell.font  = font_white_bold
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── ROWS 2-5 – Metadados da submissão ─────────────────────────────────────
    meta_rows = [
        ("Data/Hora da Submissão:", submission_dt),
        ("ID de Mobilização:",     id_mob),
        ("Solicitante:",           requester_name),
        ("E-mail:",                requester_email),
    ]
    for r_offset, (label, value) in enumerate(meta_rows, start=2):
        label_cell = ws.cell(row=r_offset, column=1, value=label)
        label_cell.fill  = fill_label
        label_cell.font  = font_label
        label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        label_cell.border = thin_border

        value_cell = ws.cell(row=r_offset, column=2, value=value)
        value_cell.fill  = fill_light
        value_cell.font  = font_value_bold
        value_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        value_cell.border = thin_border

        # Mesclar colunas restantes do valor
        if total_cols > 2:
            ws.merge_cells(start_row=r_offset, start_column=2,
                           end_row=r_offset,   end_column=total_cols)
        ws.row_dimensions[r_offset].height = 18

    # ── ROW 7 – Linha de separação ────────────────────────────────────────────
    ws.row_dimensions[7].height = 8

    # ── ROW 8 – Cabeçalho das colunas ─────────────────────────────────────────
    hdr_row = 8
    headers = ["ID Elemento", "Status"] + display_cols
    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=hdr_row, column=col_idx, value=header_text)
        cell.fill      = fill_col_hdr
        cell.font      = font_col_hdr
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border
    ws.row_dimensions[hdr_row].height = 22

    # ── ROWS 8+ – Dados ───────────────────────────────────────────────────────
    for row_offset, item in enumerate(items, start=hdr_row + 1):
        is_error = "Erro" in str(item.get("status", ""))
        row_fill = fill_error if is_error else fill_success

        # ID Elemento (ID SP)
        id_sp_cell = ws.cell(row=row_offset, column=1, value=item.get("id_sp", ""))
        id_sp_cell.fill      = row_fill
        id_sp_cell.font      = font_value_bold
        id_sp_cell.alignment = Alignment(horizontal="center", vertical="center")
        id_sp_cell.border    = thin_border

        # Status
        status_cell = ws.cell(row=row_offset, column=2, value=item.get("status", ""))
        status_cell.fill      = row_fill
        status_cell.font      = font_value
        status_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        status_cell.border    = thin_border

        # Campos do item
        fields = item.get("fields") or {}
        for col_offset, col_name in enumerate(display_cols, start=3):
            val = fields.get(col_name, "")
            data_cell = ws.cell(row=row_offset, column=col_offset, value=val)
            data_cell.fill      = row_fill
            data_cell.font      = font_value
            data_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
            data_cell.border    = thin_border

        ws.row_dimensions[row_offset].height = 18

    # ── Largura das colunas ───────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 10   # ID SP
    ws.column_dimensions["B"].width = 30   # Status
    for col_offset, col_name in enumerate(display_cols, start=3):
        estimated = max(14, min(len(col_name) + 4, 40))
        ws.column_dimensions[get_column_letter(col_offset)].width = estimated

    # ── Congelar cabeçalho ────────────────────────────────────────────────────
    ws.freeze_panes = f"A{hdr_row + 1}"

    # ── Salvar ────────────────────────────────────────────────────────────────
    os.makedirs(REPORTS_FOLDER, exist_ok=True)
    safe_id = re.sub(r"[^A-Za-z0-9]", "_", id_mob) if id_mob else "sem_id"
    # Microssegundos evitam colisão quando múltiplos grupos finalizam no mesmo segundo.
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    report_filename = f"Relatorio_Mob_{safe_id}_{timestamp}.xlsx"
    report_path     = os.path.abspath(os.path.join(REPORTS_FOLDER, report_filename))
    wb.save(report_path)
    return report_filename, report_path


def build_powershell_command(script_path, file_path, sheet_name):
    """Executa scripts PowerShell com stdout em UTF-8 para preservar acentuação."""
    utf8_preamble = (
        "[Console]::InputEncoding = [System.Text.UTF8Encoding]::new($false); "
        "[Console]::OutputEncoding = [System.Text.UTF8Encoding]::new($false); "
        "$OutputEncoding = [Console]::OutputEncoding; "
        "chcp 65001 > $null; "
    )
    escaped_script_path = script_path.replace("'", "''")
    escaped_file_path = file_path.replace("'", "''")
    escaped_sheet_name = sheet_name.replace("'", "''")

    return [
        "powershell.exe",
        "-ExecutionPolicy", "Bypass",
        "-NoProfile",
        "-Command",
        (
            f"{utf8_preamble}& '{escaped_script_path}' "
            f"-ExcelPath '{escaped_file_path}' -SheetName '{escaped_sheet_name}'"
        )
    ]


def build_powershell_populate_command(script_path, file_path):
    """Executa Populate-SharePointList.ps1 sem o parâmetro -SheetName (processamento unificado)."""
    utf8_preamble = (
        "[Console]::InputEncoding = [System.Text.UTF8Encoding]::new($false); "
        "[Console]::OutputEncoding = [System.Text.UTF8Encoding]::new($false); "
        "$OutputEncoding = [Console]::OutputEncoding; "
        "chcp 65001 > $null; "
    )
    escaped_script_path = script_path.replace("'", "''")
    escaped_file_path = file_path.replace("'", "''")

    return [
        "powershell.exe",
        "-ExecutionPolicy", "Bypass",
        "-NoProfile",
        "-Command",
        (
            f"{utf8_preamble}& '{escaped_script_path}' "
            f"-ExcelPath '{escaped_file_path}'"
        )
    ]


def build_powershell_template_update_command(script_path, template_path):
    """Executa script PowerShell de atualização do template com stdout em UTF-8."""
    utf8_preamble = (
        "[Console]::InputEncoding = [System.Text.UTF8Encoding]::new($false); "
        "[Console]::OutputEncoding = [System.Text.UTF8Encoding]::new($false); "
        "$OutputEncoding = [Console]::OutputEncoding; "
        "chcp 65001 > $null; "
    )
    escaped_script_path = script_path.replace("'", "''")
    escaped_template_path = template_path.replace("'", "''")

    return [
        "powershell.exe",
        "-ExecutionPolicy", "Bypass",
        "-NoProfile",
        "-Command",
        (
            f"{utf8_preamble}& '{escaped_script_path}' "
            f"-TemplatePath '{escaped_template_path}'"
        )
    ]


def _popen_hidden_kwargs() -> dict:
    """Oculta janelas de terminal ao executar subprocessos no Windows."""
    if os.name != 'nt':
        return {}

    startupinfo = subprocess.STARTUPINFO()
    startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    startupinfo.wShowWindow = 0
    return {
        'startupinfo': startupinfo,
        'creationflags': subprocess.CREATE_NO_WINDOW
    }


def get_template_status():
    if not os.path.exists(TEMPLATE_STATUS_FILE):
        return {'last_updated': None}

    try:
        with open(TEMPLATE_STATUS_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
            return {'last_updated': data.get('last_updated')}
    except Exception:
        return {'last_updated': None}


def save_template_status(iso_datetime):
    payload = {'last_updated': iso_datetime}
    with open(TEMPLATE_STATUS_FILE, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def decode_powershell_output(raw_output):
    """Decodifica stdout do Windows PowerShell 5.1 preservando acentuação."""
    utf8_text = raw_output.decode('utf-8', errors='replace')
    cp1252_text = raw_output.decode('cp1252', errors='replace')
    repaired_cp1252 = repair_mojibake(cp1252_text)

    utf8_score = score_decoded_text(utf8_text)
    repaired_cp1252_score = score_decoded_text(repaired_cp1252)

    if repaired_cp1252_score < utf8_score:
        return repaired_cp1252
    return utf8_text


def repair_mojibake(text):
    """Corrige trechos UTF-8 lidos como cp1252 sem afetar texto já correto."""
    previous_text = text
    for _ in range(3):
        repaired_text = repair_mojibake_once(previous_text)
        if score_decoded_text(repaired_text) >= score_decoded_text(previous_text):
            return previous_text
        previous_text = repaired_text
    return previous_text


def repair_mojibake_once(text):
    parts = re.split(r'(\s+)', text)
    repaired_parts = []
    for part in parts:
        if not part or part.isspace() or not has_mojibake_markers(part):
            repaired_parts.append(part)
            continue

        try:
            repaired_candidate = part.encode('cp1252', errors='strict').decode('utf-8', errors='strict')
        except (UnicodeEncodeError, UnicodeDecodeError):
            repaired_parts.append(part)
            continue

        if score_decoded_text(repaired_candidate) <= score_decoded_text(part):
            repaired_parts.append(repaired_candidate)
        else:
            repaired_parts.append(part)

    return ''.join(repaired_parts)


def has_mojibake_markers(text):
    return any(marker in text for marker in ('Ã', 'Â', 'â'))


def score_decoded_text(text):
    replacement_penalty = text.count('�') * 10
    mojibake_penalty = sum(text.count(marker) for marker in ('Ã', 'Â', 'â')) * 6
    return replacement_penalty + mojibake_penalty

def _count_excel_rows(file_path, sheet_name):
    """Conta linhas com dados (excluindo cabeçalho) em uma aba do Excel usando openpyxl."""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            return 0
        ws = wb[sheet_name]
        count = 0
        for row in ws.iter_rows(min_row=2):
            if any(cell.value is not None and str(cell.value).strip() for cell in row):
                count += 1
        return count
    except Exception:
        return 0
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _value_to_exact_text(value):
    """Converte valor de célula para comparação textual exata (sem normalização)."""
    if value is None:
        return ""
    return str(value)


def _row_has_any_data(values):
    for value in values:
        if value is None:
            continue
        if str(value).strip() != "":
            return True
    return False


def _read_grouped_excel(file_path: str) -> tuple[dict | None, list[str]]:
    """
    Lê abas PESSOAS e EQUIPAMENTOS exigindo coluna A = GRUPO.
    Retorna (data, errors), onde data contém headers, linhas por grupo e ordem de grupos.
    """
    errors: list[str] = []
    required_sheets = ["PESSOAS", "EQUIPAMENTOS"]

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as exc:
        return None, [f"Não foi possível abrir o Excel: {exc}"]

    try:
        for sheet in required_sheets:
            if sheet not in wb.sheetnames:
                errors.append(f"A aba obrigatória '{sheet}' não foi encontrada.")

        if errors:
            return None, errors

        groups_order: list[str] = []
        groups_seen: set[str] = set()

        data = {
            "group_order": groups_order,
            "sheets": {
                "PESSOAS": {"headers": [], "rows": [], "rows_by_group": {}},
                "EQUIPAMENTOS": {"headers": [], "rows": [], "rows_by_group": {}},
            }
        }

        for sheet in required_sheets:
            ws = wb[sheet]
            max_col = max(1, ws.max_column)
            headers = [ws.cell(row=1, column=col).value for col in range(1, max_col + 1)]
            data["sheets"][sheet]["headers"] = headers

            header_a = "" if headers[0] is None else str(headers[0]).strip()
            if header_a != "GRUPO":
                errors.append(
                    f"A aba '{sheet}' deve ter a coluna A com cabeçalho exatamente 'GRUPO'."
                )
                continue

            duplicate_keys: dict[str, set[tuple[str, ...]]] = {}

            for row_idx in range(2, ws.max_row + 1):
                row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, max_col + 1)]
                if not _row_has_any_data(row_values):
                    continue

                group_raw = row_values[0]
                if group_raw is None or str(group_raw).strip() == "":
                    errors.append(f"Aba {sheet}, linha {row_idx}: GRUPO vazio.")
                    continue

                group_key = _value_to_exact_text(group_raw)
                signature = tuple(_value_to_exact_text(v) for v in row_values)

                duplicate_bucket = duplicate_keys.setdefault(group_key, set())
                if signature in duplicate_bucket:
                    errors.append(
                        f"Aba {sheet}, linha {row_idx}: linha duplicada dentro do GRUPO '{group_key}'."
                    )
                    continue
                duplicate_bucket.add(signature)

                row_data = {
                    "excel_row": row_idx,
                    "group": group_key,
                    "values": row_values,
                }
                data["sheets"][sheet]["rows"].append(row_data)
                data["sheets"][sheet]["rows_by_group"].setdefault(group_key, []).append(row_data)

                if group_key not in groups_seen:
                    groups_seen.add(group_key)
                    groups_order.append(group_key)

        if errors:
            return None, errors

        if not groups_order:
            return None, [
                "Nenhuma linha válida encontrada para importar no SharePoint. As abas PESSOAS e EQUIPAMENTOS estão vazias."
            ]

        return data, []
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _create_group_workbook(source_data: dict, group_key: str) -> str:
    """Cria um arquivo temporário .xlsx contendo apenas linhas do grupo informado."""
    out_wb = openpyxl.Workbook()

    # Remove aba padrão
    default_sheet = out_wb.active
    out_wb.remove(default_sheet)

    for sheet_name in ["PESSOAS", "EQUIPAMENTOS"]:
        ws = out_wb.create_sheet(title=sheet_name)
        headers = source_data["sheets"][sheet_name]["headers"]
        ws.append(headers)
        group_rows = source_data["sheets"][sheet_name]["rows_by_group"].get(group_key, [])
        for row in group_rows:
            ws.append(row["values"])

    os.makedirs(REPORTS_FOLDER, exist_ok=True)
    temp_fd, temp_path = tempfile.mkstemp(prefix="mob_group_", suffix=".xlsx", dir=REPORTS_FOLDER)
    os.close(temp_fd)
    out_wb.save(temp_path)
    out_wb.close()
    return temp_path


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/download-template')
def download_template():
    return send_from_directory(app.config['UPLOAD_FOLDER'], TEMPLATE_FILENAME, as_attachment=True)


@app.route('/template-update-status', methods=['GET'])
def template_update_status():
    return jsonify(get_template_status()), 200


@app.route('/update-template', methods=['POST'])
def update_template():
    start_job()
    try:
        template_path = os.path.abspath(os.path.join(app.config['UPLOAD_FOLDER'], TEMPLATE_FILENAME))
        if not os.path.exists(template_path):
            return jsonify({'status': 'error', 'message': 'Template não encontrado.'}), 404

        script_path = os.path.abspath('Update-ExcelTemplateChoices.ps1')
        if not os.path.exists(script_path):
            return jsonify({'status': 'error', 'message': 'Script de atualização não encontrado.'}), 500

        cmd = build_powershell_template_update_command(script_path, template_path)

        process = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            universal_newlines=False,
            **_popen_hidden_kwargs()
        )
        raw_output = process.communicate()[0]
        process.wait()
        full_output = decode_powershell_output(raw_output)

        if process.returncode != 0:
            return jsonify({
                'status': 'error',
                'message': 'Falha ao atualizar template.',
                'log': full_output
            }), 500

        updated_at = datetime.now().isoformat()
        save_template_status(updated_at)

        return jsonify({
            'status': 'success',
            'message': 'Template atualizado com sucesso.',
            'last_updated': updated_at,
            'log': full_output
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Erro ao atualizar template: {str(e)}'
        }), 500
    finally:
        end_job()

@app.route('/validate', methods=['POST'])
def validate():
    """Fase 1: valida estrutura e dados por GRUPO antes do upload."""
    start_job()
    try:
        file = request.files.get('file')

        if not file or not allowed_file(file.filename):
            return jsonify({'status': 'error', 'errors': ['Arquivo inválido ou não enviado.']}), 400

        try:
            file.stream.seek(0, os.SEEK_END)
            uploaded_size = file.stream.tell()
            file.stream.seek(0)
        except Exception:
            uploaded_size = None

        if uploaded_size == 0:
            return jsonify({'status': 'error', 'errors': ['O arquivo enviado está vazio.']}), 400

        filename = secure_filename(file.filename)
        if not filename:
            return jsonify({'status': 'error', 'errors': ['Nome de arquivo inválido.']}), 400

        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)

        if os.path.getsize(file_path) == 0:
            if os.path.exists(file_path):
                os.remove(file_path)
            return jsonify({'status': 'error', 'errors': ['O arquivo enviado está vazio.']}), 400

        grouped_data, grouping_errors = _read_grouped_excel(file_path)
        if grouping_errors:
            return jsonify({
                'status': 'failed',
                'errors': grouping_errors,
                'filename': filename,
                'total_lines': 0
            }), 200

        script_path = os.path.abspath("Validate-ExcelData.ps1")
        if not os.path.exists(script_path):
            return jsonify({
                'status': 'error',
                'errors': ['Script de validação não encontrado.'],
                'filename': filename
            }), 500

        all_errors: list[str] = []
        all_logs: list[str] = []
        total_people = 0
        total_equip = 0

        group_summary = []
        for group_key in grouped_data['group_order']:
            people_count = len(grouped_data['sheets']['PESSOAS']['rows_by_group'].get(group_key, []))
            equip_count = len(grouped_data['sheets']['EQUIPAMENTOS']['rows_by_group'].get(group_key, []))
            total_people += people_count
            total_equip += equip_count
            group_summary.append({
                'group': group_key,
                'qtd_pessoas': people_count,
                'qtd_equipamentos': equip_count
            })

            temp_group_path = _create_group_workbook(grouped_data, group_key)
            try:
                for sheet_name in ('PESSOAS', 'EQUIPAMENTOS'):
                    row_count = len(grouped_data['sheets'][sheet_name]['rows_by_group'].get(group_key, []))
                    if row_count == 0:
                        continue

                    cmd = build_powershell_command(script_path, temp_group_path, sheet_name)
                    process = subprocess.Popen(
                        cmd,
                        stdout=subprocess.PIPE,
                        stderr=subprocess.STDOUT,
                        universal_newlines=False
                    )
                    raw_output = process.communicate()[0]
                    process.wait()
                    full_output = decode_powershell_output(raw_output)
                    all_logs.append(f"\n[GRUPO {group_key} | {sheet_name}]\n{full_output}")

                    start_marker = '---VALIDATION_JSON_START---'
                    end_marker = '---VALIDATION_JSON_END---'
                    if start_marker not in full_output or end_marker not in full_output:
                        all_errors.append(
                            f"[GRUPO {group_key} | {sheet_name}] Não foi possível obter resultado estruturado da validação."
                        )
                        continue

                    start_idx = full_output.index(start_marker) + len(start_marker)
                    end_idx = full_output.index(end_marker)
                    json_str = full_output[start_idx:end_idx].strip()

                    try:
                        result = json.loads(json_str)
                    except Exception as exc:
                        all_errors.append(
                            f"[GRUPO {group_key} | {sheet_name}] JSON inválido da validação: {exc}"
                        )
                        continue

                    result_errors = result.get('errors') or []
                    if result.get('status') != 'success' or result_errors:
                        for err in result_errors:
                            all_errors.append(f"[GRUPO {group_key} | {sheet_name}] {err}")
            finally:
                try:
                    if os.path.exists(temp_group_path):
                        os.remove(temp_group_path)
                except Exception:
                    pass

        total_lines = total_people + total_equip
        response_payload = {
            'status': 'success' if not all_errors else 'failed',
            'total_lines': total_lines,
            'error_count': len(all_errors),
            'errors': all_errors,
            'filename': filename,
            'group_summary': group_summary,
            'sheet_counts': {
                'PESSOAS': total_people,
                'EQUIPAMENTOS': total_equip
            },
            'log': '\n'.join(all_logs)
        }
        return jsonify(response_payload), 200

    except Exception as e:
        return jsonify({
            'status': 'error',
            'errors': [f'Erro ao executar validação: {str(e)}']
        }), 500
    finally:
        end_job()


@app.route('/run-script', methods=['POST'])
def run_script():
    """Fase 2: Upload para SharePoint por GRUPO (processamento em blocos)."""
    data = request.get_json()
    if not data:
        return Response("Erro: Dados não enviados.", status=400)

    filename = data.get('filename', '')
    if not filename or not allowed_file(filename):
        return Response("Erro: Arquivo inválido.", status=400)

    file_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(filename))
    if not os.path.exists(file_path):
        return Response("Erro: Arquivo não encontrado. Execute a validação primeiro.", status=400)

    script_path = os.path.abspath("Populate-SharePointList.ps1")
    if not os.path.exists(script_path):
        return Response("Erro: Script Populate-SharePointList.ps1 não encontrado.", status=500)

    grouped_data, grouping_errors = _read_grouped_excel(file_path)
    if grouping_errors:
        return Response("\n".join(grouping_errors), status=400, content_type='text/plain; charset=utf-8')

    def generate():
        start_job()
        consolidated_results: list[dict] = []

        try:
            groups = grouped_data['group_order']
            yield f"Arquivo recebido: {filename}\n"
            yield f"Processando {len(groups)} grupo(s) (PESSOAS + EQUIPAMENTOS)\n"
            yield f"{'-'*30}\n"

            stop_processing = False

            for idx, group_key in enumerate(groups, start=1):
                if stop_processing:
                    break

                people_count = len(grouped_data['sheets']['PESSOAS']['rows_by_group'].get(group_key, []))
                equip_count = len(grouped_data['sheets']['EQUIPAMENTOS']['rows_by_group'].get(group_key, []))

                yield f"\n---GROUP_PROGRESS:{idx}/{len(groups)}:{group_key}---\n"
                yield f"[GRUPO {group_key}] Iniciando processamento ({people_count} PESSOAS, {equip_count} EQUIPAMENTOS).\n"

                temp_group_path = _create_group_workbook(grouped_data, group_key)
                report_filename = None
                group_id_mob = ""
                group_ok = False

                try:
                    cmd = build_powershell_populate_command(script_path, temp_group_path)
                    process = subprocess.Popen(
                        cmd,
                        stdout=subprocess.PIPE,
                        stderr=subprocess.STDOUT,
                        bufsize=1,
                        universal_newlines=False
                    )

                    detected_error_in_output = False
                    error_markers = [
                        "FALHA CRÍTICA",
                        "UPLOAD CANCELADO",
                        "--- RESULT: ERROR ---",
                        "Write-Error",
                        "Erro ao adicionar item",
                        "Não foi possível gerar um ID_Mobilizacao único",
                    ]

                    in_report_json = False
                    report_json_lines: list[str] = []
                    report_data = None

                    REPORT_START = "---REPORT_JSON_START---"
                    REPORT_END = "---REPORT_JSON_END---"

                    while True:
                        chunk = process.stdout.readline()
                        if not chunk:
                            break
                        decoded = decode_powershell_output(chunk)
                        stripped = decoded.strip()

                        if stripped == REPORT_START:
                            in_report_json = True
                            continue
                        if stripped == REPORT_END:
                            in_report_json = False
                            continue
                        if in_report_json:
                            report_json_lines.append(stripped)
                            continue

                        if any(marker in decoded for marker in error_markers):
                            detected_error_in_output = True

                        yield decoded

                    process.wait()

                    if report_json_lines:
                        try:
                            report_data = json.loads("".join(report_json_lines))
                            group_id_mob = str(report_data.get('id_mobilizacao', '') or '')
                            report_filename, _ = generate_report_excel(report_data)
                            yield f"---REPORT_FILE:{report_filename}---\n"
                        except Exception as exc:
                            yield f"[AVISO] Não foi possível gerar o relatório Excel do grupo {group_key}: {exc}\n"

                    group_ok = process.returncode == 0 and not detected_error_in_output
                    status_text = "success" if group_ok else "error"
                    group_result = {
                        'id_mobilizacao': group_id_mob,
                        'status': status_text,
                        'qtd_pessoas': people_count,
                        'qtd_equipamentos': equip_count,
                        'report_filename': report_filename
                    }
                    consolidated_results.append(group_result)
                    yield f"---GROUP_RESULT:{json.dumps(group_result, ensure_ascii=False)}---\n"

                    if group_ok:
                        yield f"[GRUPO {group_key}] Concluído com sucesso.\n"
                    else:
                        yield f"[GRUPO {group_key}] Falhou. Encerrando processamento dos próximos grupos.\n"
                        stop_processing = True

                except Exception as exc:
                    group_result = {
                        'id_mobilizacao': group_id_mob,
                        'status': 'error',
                        'qtd_pessoas': people_count,
                        'qtd_equipamentos': equip_count,
                        'report_filename': report_filename
                    }
                    consolidated_results.append(group_result)
                    yield f"---GROUP_RESULT:{json.dumps(group_result, ensure_ascii=False)}---\n"
                    yield f"[ERRO DE EXECUÇÃO][GRUPO {group_key}]: {str(exc)}\n"
                    stop_processing = True
                finally:
                    try:
                        if os.path.exists(temp_group_path):
                            os.remove(temp_group_path)
                    except Exception:
                        pass

            yield "---GROUP_SUMMARY_JSON_START---\n"
            yield json.dumps({'groups': consolidated_results}, ensure_ascii=False)
            yield "\n---GROUP_SUMMARY_JSON_END---\n"

            all_ok = consolidated_results and all(g.get('status') == 'success' for g in consolidated_results)
            if all_ok:
                yield f"\n{'-'*30}\n[SUCESSO] Todos os grupos foram processados com sucesso.\n"
            else:
                yield f"\n{'-'*30}\n[ERRO] Processamento encerrado com falha em um grupo.\n"

        except Exception as e:
            yield f"\n[ERRO DE EXECUÇÃO]: {str(e)}\n"
        finally:
            end_job()

    return Response(stream_with_context(generate()), content_type='text/plain; charset=utf-8')

@app.route('/shutdown', methods=['POST'])
def shutdown():
    """Encerra o servidor Flask."""
    def _shutdown():
        os._exit(0)
    t = threading.Timer(0.5, _shutdown)
    t.daemon = True
    t.start()
    return jsonify({'message': 'Servidor encerrado.'}), 200


@app.route('/list-reports')
def list_reports():
    """Lista os relatórios Excel gerados, ordenados do mais recente para o mais antigo."""
    reports_abs = os.path.abspath(REPORTS_FOLDER)
    if not os.path.isdir(reports_abs):
        return jsonify({'reports': []})
    files = []
    for fname in os.listdir(reports_abs):
        if fname.lower().endswith('.xlsx'):
            fpath = os.path.join(reports_abs, fname)
            try:
                mtime = os.path.getmtime(fpath)
                size  = os.path.getsize(fpath)
            except OSError:
                continue
            files.append({'name': fname, 'mtime': mtime, 'size': size})
    files.sort(key=lambda f: f['mtime'], reverse=True)
    for f in files:
        from datetime import datetime as _dt
        f['modified'] = _dt.fromtimestamp(f['mtime']).strftime('%d/%m/%Y %H:%M:%S')
        del f['mtime']
    return jsonify({'reports': files})


@app.route('/download-report/<path:filename>')
def download_report(filename):
    """Serve o relatório Excel gerado após a submissão."""
    safe_name = secure_filename(filename)
    reports_abs = os.path.abspath(REPORTS_FOLDER)
    file_abs    = os.path.abspath(os.path.join(reports_abs, safe_name))
    # Previne path traversal
    if not file_abs.startswith(reports_abs + os.sep):
        return jsonify({'error': 'Acesso negado.'}), 403
    if not os.path.exists(file_abs):
        return jsonify({'error': 'Relatório não encontrado.'}), 404
    return send_from_directory(reports_abs, safe_name, as_attachment=True)


if __name__ == '__main__':
    print("Servidor rodando em http://localhost:5000")
    app.run(debug=False, host='0.0.0.0', port=5000, threaded=True)
